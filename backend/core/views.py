from datetime import date, datetime, timedelta

from django.contrib.auth import authenticate, login, logout
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from .forms import (
    InvitadoFrecuenteForm,
    ReclamoForm,
    ReservaSUMForm,
    SolicitudModificacionFamiliaForm,
    VisitaForm,
    VisitaEspontaneaForm,
    EncomiendaForm,
    EntregaEncomiendaForm,
)

from .models import (
    ContactoUtil,
    Documento,
    Encomienda,
    Integrante,
    InvitadoFrecuente,
    Noticia,
    Reclamo,
    ReservaSUM,
    SolicitudModificacionFamilia,
    Visita,
)


# -------------------------------------------------
# HOME Y LOGIN
# -------------------------------------------------

def home(request):
    return render(request, "core/home.html")


def login_view(request):
    if request.method == "POST":
        numero_lote = request.POST.get("numero_lote")
        password = request.POST.get("password")

        usuario = authenticate(
            request,
            username=numero_lote,
            password=password
        )

        if usuario is not None:
            login(request, usuario)

            if usuario.is_staff:
                return redirect("seguridad_dashboard")

            return redirect("portal")

        return render(
            request,
            "core/home.html",
            {
                "error": "Número de lote o contraseña incorrectos."
            }
        )

    return render(request, "core/home.html")


# -------------------------------------------------
# CUMPLEAÑOS - FUNCIÓN AUXILIAR
# -------------------------------------------------

def obtener_proximos_cumpleanios(limite=None):
    hoy = timezone.localdate()

    integrantes = Integrante.objects.filter(
        activo=True
    )

    cumpleanios = []

    for integrante in integrantes:
        nacimiento = integrante.fecha_nacimiento

        try:
            proximo = date(
                hoy.year,
                nacimiento.month,
                nacimiento.day
            )

        except ValueError:
            proximo = date(
                hoy.year,
                2,
                28
            )

        if proximo < hoy:

            try:
                proximo = date(
                    hoy.year + 1,
                    nacimiento.month,
                    nacimiento.day
                )

            except ValueError:
                proximo = date(
                    hoy.year + 1,
                    2,
                    28
                )

        dias_faltantes = (
            proximo - hoy
        ).days

        cumpleanios.append(
            {
                "integrante": integrante,
                "fecha": proximo,
                "dias_faltantes": dias_faltantes,
            }
        )

    cumpleanios.sort(
        key=lambda item: item["fecha"]
    )

    if limite:
        cumpleanios = cumpleanios[:limite]

    return cumpleanios


# -------------------------------------------------
# PORTAL
# -------------------------------------------------

def portal(request):
    if not request.user.is_authenticated:
        return redirect("home")

    if request.user.is_staff:
        return redirect("seguridad_dashboard")

    lote = request.user.lote
    hoy = timezone.localdate()

    # NOTICIAS

    noticias = Noticia.objects.filter(
        activa=True
    ).order_by(
        "-fecha_publicacion"
    )[:3]

    # RESERVAS DEL SUM

    proximas_reservas = ReservaSUM.objects.filter(
        lote=lote
    ).order_by(
        "-fecha",
        "-fecha_creacion"
    )[:3]

    for reserva in proximas_reservas:

        fecha_limite_cancelacion = (
            reserva.fecha - timedelta(days=7)
        )

        reserva.puede_cancelar = (
            hoy <= fecha_limite_cancelacion
            and reserva.estado != "cancelada"
        )

    # CUMPLEAÑOS DEL MES

    cumpleanios_mes = []

    integrantes_activos = Integrante.objects.filter(
        activo=True
    )

    for integrante in integrantes_activos:

        if integrante.fecha_nacimiento.month == hoy.month:
            cumpleanios_mes.append(integrante)

    cumpleanios_mes.sort(
        key=lambda integrante:
        integrante.fecha_nacimiento.day
    )

    # ENCOMIENDAS PENDIENTES DEL LOTE

    encomiendas_pendientes = Encomienda.objects.filter(
        lote=lote,
        estado="pendiente"
    ).order_by(
        "-fecha_recepcion"
    )

    cantidad_encomiendas_pendientes = (
        encomiendas_pendientes.count()
    )

    ultimas_encomiendas_pendientes = (
        encomiendas_pendientes[:3]
    )

    return render(
        request,
        "core/portal.html",
        {
            "lote": lote,
            "noticias": noticias,
            "proximas_reservas": proximas_reservas,
            "cumpleanios_mes": cumpleanios_mes,
            "cantidad_encomiendas_pendientes":
                cantidad_encomiendas_pendientes,
            "ultimas_encomiendas_pendientes":
                ultimas_encomiendas_pendientes,
        }
    )


# -------------------------------------------------
# NOTICIAS
# -------------------------------------------------

def noticias_view(request):
    if not request.user.is_authenticated:
        return redirect("home")

    noticias = Noticia.objects.filter(
        activa=True
    ).order_by(
        "-fecha_publicacion"
    )

    return render(
        request,
        "core/noticias.html",
        {
            "noticias": noticias
        }
    )


# -------------------------------------------------
# CUMPLEAÑOS
# -------------------------------------------------

def cumpleanios_view(request):
    if not request.user.is_authenticated:
        return redirect("home")

    cumpleanios = obtener_proximos_cumpleanios()

    return render(
        request,
        "core/cumpleanios.html",
        {
            "cumpleanios": cumpleanios,
        }
    )


# -------------------------------------------------
# RESERVAS SUM
# -------------------------------------------------

def mis_reservas_sum(request):
    if not request.user.is_authenticated:
        return redirect("home")

    lote = request.user.lote
    hoy = timezone.localdate()

    reservas = ReservaSUM.objects.filter(
        lote=lote
    ).order_by(
        "-fecha",
        "-fecha_creacion"
    )

    for reserva in reservas:

        fecha_limite_cancelacion = (
            reserva.fecha - timedelta(days=7)
        )

        reserva.puede_cancelar = (
            hoy <= fecha_limite_cancelacion
            and reserva.estado != "cancelada"
        )

    return render(
        request,
        "core/mis_reservas_sum.html",
        {
            "lote": lote,
            "reservas": reservas,
        }
    )


def disponibilidad_sum(request):
    if not request.user.is_authenticated:
        return redirect("home")

    lote = request.user.lote

    fecha_texto = request.GET.get("fecha")

    fecha_consultada = None
    turno_dia_disponible = None
    turno_noche_disponible = None

    if fecha_texto:

        try:

            fecha_consultada = datetime.strptime(
                fecha_texto,
                "%Y-%m-%d"
            ).date()

            turno_dia_ocupado = ReservaSUM.objects.filter(
                fecha=fecha_consultada,
                turno="dia"
            ).exclude(
                estado="cancelada"
            ).exists()

            turno_noche_ocupado = ReservaSUM.objects.filter(
                fecha=fecha_consultada,
                turno="noche"
            ).exclude(
                estado="cancelada"
            ).exists()

            turno_dia_disponible = (
                not turno_dia_ocupado
            )

            turno_noche_disponible = (
                not turno_noche_ocupado
            )

        except ValueError:
            fecha_consultada = None

    return render(
        request,
        "core/disponibilidad_sum.html",
        {
            "lote": lote,
            "fecha_consultada": fecha_consultada,
            "fecha_texto": fecha_texto,
            "turno_dia_disponible": turno_dia_disponible,
            "turno_noche_disponible": turno_noche_disponible,
        }
    )


def reservar_sum(request):
    if not request.user.is_authenticated:
        return redirect("home")

    lote = request.user.lote

    fecha_inicial = request.GET.get("fecha")
    turno_inicial = request.GET.get("turno")

    if request.method == "POST":

        form = ReservaSUMForm(
            request.POST
        )

        if form.is_valid():

            reserva = form.save(
                commit=False
            )

            reserva.lote = lote
            reserva.estado = "pendiente"

            try:

                reserva.save()

                return render(
                    request,
                    "core/reserva_sum_ok.html",
                    {
                        "reserva": reserva
                    }
                )

            except Exception:

                form.add_error(
                    None,
                    (
                        "Ese turno ya está reservado "
                        "para la fecha seleccionada."
                    )
                )

    else:

        form = ReservaSUMForm(
            initial={
                "fecha": fecha_inicial,
                "turno": turno_inicial,
            }
        )

    return render(
        request,
        "core/reservar_sum.html",
        {
            "form": form,
            "lote": lote,
        }
    )


def cancelar_reserva_sum(
    request,
    reserva_id
):
    if not request.user.is_authenticated:
        return redirect("home")

    lote = request.user.lote

    reserva = get_object_or_404(
        ReservaSUM,
        id=reserva_id,
        lote=lote
    )

    if request.method == "POST":

        hoy = timezone.localdate()

        fecha_limite = (
            reserva.fecha - timedelta(days=7)
        )

        if (
            hoy <= fecha_limite
            and reserva.estado != "cancelada"
        ):

            reserva.estado = "cancelada"
            reserva.save()

    return redirect("portal")


# -------------------------------------------------
# MI FAMILIA
# -------------------------------------------------

def mi_familia(request):
    if not request.user.is_authenticated:
        return redirect("home")

    lote = request.user.lote

    integrantes = Integrante.objects.filter(
        lote=lote,
        activo=True
    ).order_by(
        "apellido",
        "nombre"
    )

    solicitudes = (
        SolicitudModificacionFamilia.objects
        .filter(
            lote=lote
        )
        .order_by(
            "-fecha_creacion"
        )
    )

    return render(
        request,
        "core/mi_familia.html",
        {
            "lote": lote,
            "integrantes": integrantes,
            "solicitudes": solicitudes,
        }
    )


def solicitar_modificacion_familia(request):
    if not request.user.is_authenticated:
        return redirect("home")

    lote = request.user.lote

    if request.method == "POST":

        form = SolicitudModificacionFamiliaForm(
            request.POST,
            lote=lote
        )

        if form.is_valid():

            solicitud = form.save(
                commit=False
            )

            solicitud.lote = lote
            solicitud.estado = "pendiente"

            solicitud.save()

            return redirect(
                "mi_familia"
            )

    else:

        form = SolicitudModificacionFamiliaForm(
            lote=lote
        )

    return render(
        request,
        "core/solicitar_modificacion_familia.html",
        {
            "form": form,
            "lote": lote,
        }
    )


# -------------------------------------------------
# RECLAMOS
# -------------------------------------------------

def mis_reclamos(request):
    if not request.user.is_authenticated:
        return redirect("home")

    lote = request.user.lote

    reclamos = Reclamo.objects.filter(
        lote=lote
    ).order_by(
        "-fecha_creacion"
    )

    return render(
        request,
        "core/mis_reclamos.html",
        {
            "lote": lote,
            "reclamos": reclamos,
        }
    )


def nuevo_reclamo(request):
    if not request.user.is_authenticated:
        return redirect("home")

    lote = request.user.lote

    if request.method == "POST":

        form = ReclamoForm(
            request.POST
        )

        if form.is_valid():

            reclamo = form.save(
                commit=False
            )

            reclamo.lote = lote
            reclamo.estado = "pendiente"

            reclamo.save()

            return redirect(
                "mis_reclamos"
            )

    else:

        form = ReclamoForm()

    return render(
        request,
        "core/nuevo_reclamo.html",
        {
            "form": form,
            "lote": lote,
        }
    )


# -------------------------------------------------
# ENCOMIENDAS
# -------------------------------------------------

def mis_encomiendas(request):
    if not request.user.is_authenticated:
        return redirect("home")

    lote = request.user.lote

    encomiendas = Encomienda.objects.filter(
        lote=lote
    ).order_by(
        "-fecha_recepcion"
    )

    pendientes = encomiendas.filter(
        estado="pendiente"
    )

    cantidad_pendientes = pendientes.count()

    return render(
        request,
        "core/mis_encomiendas.html",
        {
            "lote": lote,
            "encomiendas": encomiendas,
            "cantidad_pendientes": cantidad_pendientes,
        }
    )


# -------------------------------------------------
# DOCUMENTOS
# -------------------------------------------------

def documentos_view(request):
    if not request.user.is_authenticated:
        return redirect("home")

    documentos = Documento.objects.filter(
        activo=True
    ).order_by(
        "categoria",
        "-fecha_publicacion"
    )

    categorias = []

    for codigo, nombre in Documento.CATEGORIAS:

        documentos_categoria = documentos.filter(
            categoria=codigo
        )

        if documentos_categoria.exists():
            categorias.append(
                {
                    "codigo": codigo,
                    "nombre": nombre,
                    "documentos": documentos_categoria,
                }
            )

    return render(
        request,
        "core/documentos.html",
        {
            "documentos": documentos,
            "categorias": categorias,
        }
    )


# -------------------------------------------------
# CONTACTOS ÚTILES
# -------------------------------------------------

def contactos_utiles(request):
    if not request.user.is_authenticated:
        return redirect("home")

    contactos = ContactoUtil.objects.filter(
        activo=True
    ).order_by(
        "orden",
        "categoria",
        "nombre"
    )

    categorias = []

    for codigo, nombre in ContactoUtil.CATEGORIAS:

        contactos_categoria = contactos.filter(
            categoria=codigo
        )

        if contactos_categoria.exists():
            categorias.append(
                {
                    "codigo": codigo,
                    "nombre": nombre,
                    "contactos": contactos_categoria,
                }
            )

    return render(
        request,
        "core/contactos_utiles.html",
        {
            "categorias": categorias,
        }
    )

    # -------------------------------------------------
# VISITAS
# -------------------------------------------------

def visitas_view(request):
    if not request.user.is_authenticated:
        return redirect("home")

    lote = request.user.lote

    proximas_visitas = Visita.objects.filter(
        lote=lote,
        fecha__gte=timezone.localdate(),
        estado="autorizada",
    ).order_by(
        "fecha",
        "apellido",
        "nombre",
    )[:10]

    return render(
        request,
        "core/visitas.html",
        {
            "lote": lote,
            "proximas_visitas": proximas_visitas,
        }
    )


def autorizar_visita(request):
    if not request.user.is_authenticated:
        return redirect("home")

    lote = request.user.lote

    if request.method == "POST":

        form = VisitaForm(
            request.POST,
            lote=lote,
        )

        if form.is_valid():

            visita = form.save(
                commit=False
            )

            visita.lote = lote
            visita.estado = "autorizada"

            if visita.invitado:

                visita.nombre = visita.invitado.nombre
                visita.apellido = visita.invitado.apellido
                visita.dni = visita.invitado.dni

                if not visita.patente:
                    visita.patente = visita.invitado.patente

            visita.save()

            return redirect(
                "historial_visitas"
            )

    else:

        form = VisitaForm(
            lote=lote
        )

    return render(
        request,
        "core/autorizar_visita.html",
        {
            "form": form,
            "lote": lote,
        }
    )


def historial_visitas(request):
    if not request.user.is_authenticated:
        return redirect("home")

    lote = request.user.lote

    visitas = Visita.objects.filter(
        lote=lote
    ).order_by(
        "-fecha",
        "-fecha_creacion",
    )

    return render(
        request,
        "core/historial_visitas.html",
        {
            "lote": lote,
            "visitas": visitas,
        }
    )


def agenda_visitas(request):
    if not request.user.is_authenticated:
        return redirect("home")

    lote = request.user.lote
    hoy = timezone.localdate()

    visitas = Visita.objects.filter(
        lote=lote,
        fecha__gte=hoy,
        estado="autorizada",
    ).order_by(
        "fecha",
        "apellido",
        "nombre",
    )

    return render(
        request,
        "core/agenda_visitas.html",
        {
            "lote": lote,
            "visitas": visitas,
            "hoy": hoy,
        }
    )


def invitados_frecuentes(request):
    if not request.user.is_authenticated:
        return redirect("home")

    lote = request.user.lote

    invitados = InvitadoFrecuente.objects.filter(
        lote=lote,
        activo=True,
    ).order_by(
        "apellido",
        "nombre",
    )

    return render(
        request,
        "core/invitados_frecuentes.html",
        {
            "lote": lote,
            "invitados": invitados,
        }
    )


def nuevo_invitado_frecuente(request):
    if not request.user.is_authenticated:
        return redirect("home")

    lote = request.user.lote

    if request.method == "POST":

        form = InvitadoFrecuenteForm(
            request.POST
        )

        if form.is_valid():

            invitado = form.save(
                commit=False
            )

            invitado.lote = lote
            invitado.activo = True

            invitado.save()

            return redirect(
                "invitados_frecuentes"
            )

    else:

        form = InvitadoFrecuenteForm()

    return render(
        request,
        "core/nuevo_invitado_frecuente.html",
        {
            "form": form,
            "lote": lote,
        }
    )


def cancelar_visita(
    request,
    visita_id
):
    if not request.user.is_authenticated:
        return redirect("home")

    lote = request.user.lote

    visita = get_object_or_404(
        Visita,
        id=visita_id,
        lote=lote,
    )

    if request.method == "POST":

        visita.estado = "cancelada"
        visita.save()

    return redirect(
        "historial_visitas"
    )


# -------------------------------------------------
# CARGA MASIVA DE VISITAS
# -------------------------------------------------

def carga_masiva_visitas(request):
    if not request.user.is_authenticated:
        return redirect("home")

    from datetime import date as fecha_clase
    from datetime import datetime as fecha_hora

    import openpyxl

    lote = request.user.lote
    errores = []
    visitas_preview = []

    if request.method == "POST":
        archivo = request.FILES.get("archivo")

        if not archivo:
            errores.append(
                "Debés seleccionar un archivo Excel."
            )

        elif not archivo.name.lower().endswith(
            (".xlsx", ".xlsm")
        ):
            errores.append(
                "El archivo debe estar en formato Excel (.xlsx)."
            )

        else:
            try:
                workbook = openpyxl.load_workbook(
                    archivo,
                    data_only=True
                )

                if "Visitas" in workbook.sheetnames:
                    hoja = workbook["Visitas"]
                else:
                    hoja = workbook.active

                encabezados_esperados = [
                    "Nombre",
                    "Apellido",
                    "DNI",
                    "Patente",
                    "Fecha",
                    "Evento",
                    "Observaciones",
                ]

                encabezados = []

                for celda in hoja[1]:
                    valor = celda.value

                    if valor is None:
                        encabezados.append("")
                    else:
                        encabezados.append(
                            str(valor).strip()
                        )

                if encabezados[:7] != encabezados_esperados:
                    errores.append(
                        "La estructura del archivo no es correcta. "
                        "Descargá y utilizá la plantilla oficial."
                    )

                else:
                    for numero_fila, fila in enumerate(
                        hoja.iter_rows(
                            min_row=2,
                            values_only=True
                        ),
                        start=2
                    ):
                        valores = list(fila[:7])

                        while len(valores) < 7:
                            valores.append(None)

                        (
                            nombre,
                            apellido,
                            dni,
                            patente,
                            fecha_visita,
                            evento,
                            observaciones,
                        ) = valores

                        fila_vacia = not any(
                            valor not in (None, "")
                            for valor in valores
                        )

                        if fila_vacia:
                            continue

                        nombre = (
                            str(nombre).strip()
                            if nombre is not None
                            else ""
                        )

                        apellido = (
                            str(apellido).strip()
                            if apellido is not None
                            else ""
                        )

                        dni = (
                            str(dni).strip()
                            if dni is not None
                            else ""
                        )

                        if dni.endswith(".0"):
                            dni = dni[:-2]

                        patente = (
                            str(patente).strip().upper()
                            if patente is not None
                            else ""
                        )

                        evento = (
                            str(evento).strip()
                            if evento is not None
                            else ""
                        )

                        observaciones = (
                            str(observaciones).strip()
                            if observaciones is not None
                            else ""
                        )

                        errores_fila = []

                        if not nombre:
                            errores_fila.append(
                                "Falta el nombre."
                            )

                        if not apellido:
                            errores_fila.append(
                                "Falta el apellido."
                            )

                        if not dni:
                            errores_fila.append(
                                "Falta el DNI."
                            )

                        fecha_convertida = None

                        if isinstance(
                            fecha_visita,
                            fecha_hora
                        ):
                            fecha_convertida = fecha_visita.date()

                        elif isinstance(
                            fecha_visita,
                            fecha_clase
                        ):
                            fecha_convertida = fecha_visita

                        elif fecha_visita:
                            texto_fecha = str(
                                fecha_visita
                            ).strip()

                            formatos = [
                                "%d/%m/%Y",
                                "%d-%m-%Y",
                                "%Y-%m-%d",
                            ]

                            for formato in formatos:
                                try:
                                    fecha_convertida = (
                                        fecha_hora.strptime(
                                            texto_fecha,
                                            formato
                                        ).date()
                                    )
                                    break
                                except ValueError:
                                    pass

                        if not fecha_convertida:
                            errores_fila.append(
                                "La fecha no es válida."
                            )

                        if errores_fila:
                            errores.append(
                                "Fila "
                                + str(numero_fila)
                                + ": "
                                + " ".join(errores_fila)
                            )
                            continue

                        visitas_preview.append(
                            {
                                "nombre": nombre,
                                "apellido": apellido,
                                "dni": dni,
                                "patente": patente,
                                "fecha": fecha_convertida.isoformat(),
                                "evento": evento,
                                "observaciones": observaciones,
                            }
                        )

                    if (
                        not errores
                        and not visitas_preview
                    ):
                        errores.append(
                            "El archivo no contiene visitas para importar."
                        )

                    if visitas_preview:
                        request.session[
                            "visitas_carga_masiva"
                        ] = visitas_preview

            except Exception as error:
                errores.append(
                    "No se pudo leer el archivo Excel. "
                    "Verificá que el archivo sea válido."
                )

                print(
                    "Error carga masiva:",
                    error
                )

    return render(
        request,
        "core/carga_masiva_visitas.html",
        {
            "lote": lote,
            "errores": errores,
            "visitas_preview": visitas_preview,
        }
    )


def confirmar_carga_masiva_visitas(request):
    if not request.user.is_authenticated:
        return redirect("home")

    from datetime import date as fecha_clase

    lote = request.user.lote

    if request.method != "POST":
        return redirect(
            "carga_masiva_visitas"
        )

    datos = request.session.get(
        "visitas_carga_masiva",
        []
    )

    if not datos:
        return redirect(
            "carga_masiva_visitas"
        )

    visitas = []

    for dato in datos:
        visitas.append(
            Visita(
                lote=lote,
                nombre=dato["nombre"],
                apellido=dato["apellido"],
                dni=dato["dni"],
                patente=dato.get(
                    "patente",
                    ""
                ),
                fecha=fecha_clase.fromisoformat(
                    dato["fecha"]
                ),
                evento=dato.get(
                    "evento",
                    ""
                ),
                observaciones=dato.get(
                    "observaciones",
                    ""
                ),
                estado="autorizada",
            )
        )

    Visita.objects.bulk_create(
        visitas
    )

    if "visitas_carga_masiva" in request.session:
        del request.session[
            "visitas_carga_masiva"
        ]

    return render(
        request,
        "core/carga_masiva_visitas_ok.html",
        {
            "lote": lote,
            "cantidad": len(visitas),
        }
    )


def descargar_plantilla_visitas(request):
    if not request.user.is_authenticated:
        return redirect("home")

    from io import BytesIO

    import openpyxl

    from django.http import HttpResponse
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill

    workbook = openpyxl.Workbook()

    hoja = workbook.active
    hoja.title = "Visitas"

    encabezados = [
        "Nombre",
        "Apellido",
        "DNI",
        "Patente",
        "Fecha",
        "Evento",
        "Observaciones",
    ]

    hoja.append(encabezados)

    for celda in hoja[1]:
        celda.font = Font(
            bold=True
        )

        celda.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAD3"
        )

        celda.alignment = Alignment(
            horizontal="center"
        )

    anchos = {
        "A": 20,
        "B": 20,
        "C": 18,
        "D": 15,
        "E": 15,
        "F": 30,
        "G": 40,
    }

    for columna, ancho in anchos.items():
        hoja.column_dimensions[
            columna
        ].width = ancho

    hoja.freeze_panes = "A2"

    instrucciones = workbook.create_sheet(
        "Instrucciones"
    )

    instrucciones["A1"] = "Carga Masiva de Visitas"
    instrucciones["A1"].font = Font(
        bold=True,
        size=14
    )

    instrucciones["A3"] = (
        "Completá una persona por fila "
        "en la hoja 'Visitas'."
    )

    instrucciones["A4"] = (
        "No modifiques los nombres "
        "de las columnas."
    )

    instrucciones["A5"] = (
        "Nombre, Apellido, DNI y Fecha "
        "son obligatorios."
    )

    instrucciones["A6"] = (
        "Patente, Evento y Observaciones "
        "son opcionales."
    )

    instrucciones["A7"] = (
        "La fecha puede cargarse como "
        "DD/MM/AAAA."
    )

    instrucciones["A9"] = "Ejemplo:"

    instrucciones["A10"] = "Nombre"
    instrucciones["B10"] = "Apellido"
    instrucciones["C10"] = "DNI"
    instrucciones["D10"] = "Patente"
    instrucciones["E10"] = "Fecha"
    instrucciones["F10"] = "Evento"
    instrucciones["G10"] = "Observaciones"

    instrucciones["A11"] = "Juan"
    instrucciones["B11"] = "Pérez"
    instrucciones["C11"] = "30123456"
    instrucciones["D11"] = "AB123CD"
    instrucciones["E11"] = "15/09/2026"
    instrucciones["F11"] = "Cumpleaños"
    instrucciones["G11"] = "Invitado de ejemplo"

    for columna, ancho in anchos.items():
        instrucciones.column_dimensions[
            columna
        ].width = ancho

    archivo = BytesIO()
    workbook.save(archivo)
    archivo.seek(0)

    respuesta = HttpResponse(
        archivo.getvalue(),
        content_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    respuesta[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="Plantilla_Carga_Visitas.xlsx"'
    )

    return respuesta




# -------------------------------------------------
# SEGURIDAD / PORTERÍA - DASHBOARD
# -------------------------------------------------

def seguridad_dashboard(request):
    if not request.user.is_authenticated:
        return redirect("home")

    if not request.user.is_staff:
        return redirect("portal")

    hoy = timezone.localdate()

    personas_dentro = Visita.objects.filter(
        estado="autorizada",
        fecha_hora_ingreso__isnull=False,
        fecha_hora_salida__isnull=True,
    ).count()

    visitas_hoy_pendientes = Visita.objects.filter(
        estado="autorizada",
        fecha=hoy,
        fecha_hora_ingreso__isnull=True,
    ).count()

    proximas_visitas = Visita.objects.filter(
        estado="autorizada",
        fecha__gt=hoy,
        fecha_hora_ingreso__isnull=True,
    ).count()

    encomiendas_pendientes = Encomienda.objects.filter(
        estado="pendiente"
    ).count()

    reservas_sum_hoy = ReservaSUM.objects.filter(
        fecha=hoy
    ).exclude(
        estado="cancelada"
    ).select_related(
        "lote"
    ).order_by("turno")

    return render(
        request,
        "core/seguridad_dashboard.html",
        {
            "hoy": hoy,
            "personas_dentro": personas_dentro,
            "visitas_hoy_pendientes": visitas_hoy_pendientes,
            "proximas_visitas": proximas_visitas,
            "encomiendas_pendientes": encomiendas_pendientes,
            "reservas_sum_hoy": reservas_sum_hoy,
        }
    )


# -------------------------------------------------
# SEGURIDAD / PORTERÍA - CONSULTA DE VISITAS
# -------------------------------------------------

def seguridad_visitas(request):
    from django.db.models import Q

    if not request.user.is_authenticated:
        return redirect("home")

    if not request.user.is_staff:
        return redirect("portal")

    hoy = timezone.localdate()
    busqueda = request.GET.get("q", "").strip()

    personas_dentro = Visita.objects.filter(
        estado="autorizada",
        fecha_hora_ingreso__isnull=False,
        fecha_hora_salida__isnull=True,
    ).select_related(
        "lote"
    ).order_by(
        "fecha_hora_ingreso"
    )

    visitas_hoy = Visita.objects.filter(
        estado="autorizada",
        fecha=hoy,
        fecha_hora_ingreso__isnull=True,
    ).select_related(
        "lote"
    ).order_by(
        "apellido",
        "nombre",
    )

    proximas_visitas = Visita.objects.filter(
        estado="autorizada",
        fecha__gt=hoy,
        fecha_hora_ingreso__isnull=True,
    ).select_related(
        "lote"
    ).order_by(
        "fecha",
        "apellido",
        "nombre",
    )

    if busqueda:
        filtro = (
            Q(dni__icontains=busqueda)
            | Q(patente__icontains=busqueda)
            | Q(apellido__icontains=busqueda)
        )

        personas_dentro = personas_dentro.filter(filtro)
        visitas_hoy = visitas_hoy.filter(filtro)
        proximas_visitas = proximas_visitas.filter(filtro)

    proximas_visitas = proximas_visitas[:100]

    return render(
        request,
        "core/seguridad_visitas.html",
        {
            "personas_dentro": personas_dentro,
            "visitas_hoy": visitas_hoy,
            "proximas_visitas": proximas_visitas,
            "busqueda": busqueda,
            "hoy": hoy,
        }
    )


# -------------------------------------------------
# SEGURIDAD / PORTERÍA - HISTORIAL
# -------------------------------------------------

def historial_seguridad(request):
    from django.db.models import Q

    if not request.user.is_authenticated:
        return redirect("home")

    if not request.user.is_staff:
        return redirect("portal")

    busqueda = request.GET.get("q", "").strip()
    fecha_desde = request.GET.get("desde", "").strip()
    fecha_hasta = request.GET.get("hasta", "").strip()

    movimientos = Visita.objects.filter(
        Q(fecha_hora_ingreso__isnull=False)
        | Q(fecha_hora_salida__isnull=False)
    ).select_related("lote")

    if busqueda:
        movimientos = movimientos.filter(
            Q(dni__icontains=busqueda)
            | Q(patente__icontains=busqueda)
            | Q(apellido__icontains=busqueda)
            | Q(nombre__icontains=busqueda)
            | Q(lote__numero__icontains=busqueda)
        )

    if fecha_desde:
        movimientos = movimientos.filter(
            fecha_hora_ingreso__date__gte=fecha_desde
        )

    if fecha_hasta:
        movimientos = movimientos.filter(
            fecha_hora_ingreso__date__lte=fecha_hasta
        )

    movimientos = movimientos.order_by(
        "-fecha_hora_ingreso",
        "-fecha",
    )[:500]

    return render(
        request,
        "core/historial_seguridad.html",
        {
            "movimientos": movimientos,
            "busqueda": busqueda,
            "fecha_desde": fecha_desde,
            "fecha_hasta": fecha_hasta,
        }
    )


# -------------------------------------------------
# SEGURIDAD / PORTERÍA - VISITA ESPONTÁNEA
# -------------------------------------------------

def visita_espontanea(request):
    if not request.user.is_authenticated:
        return redirect("home")

    if not request.user.is_staff:
        return redirect("portal")

    if request.method == "POST":
        form = VisitaEspontaneaForm(request.POST)

        if form.is_valid():
            visita = form.save(commit=False)
            visita.fecha = timezone.localdate()
            visita.estado = "autorizada"
            visita.fecha_hora_ingreso = timezone.now()
            visita.save()

            return redirect("seguridad_visitas")

    else:
        form = VisitaEspontaneaForm()

    return render(
        request,
        "core/visita_espontanea.html",
        {
            "form": form,
        }
    )


# -------------------------------------------------
# SEGURIDAD / PORTERÍA - ENCOMIENDAS
# -------------------------------------------------

def seguridad_encomiendas(request):
    from django.db.models import Q

    if not request.user.is_authenticated:
        return redirect("home")

    if not request.user.is_staff:
        return redirect("portal")

    busqueda = request.GET.get("q", "").strip()

    pendientes = Encomienda.objects.filter(
        estado="pendiente"
    ).select_related(
        "lote"
    ).order_by(
        "-fecha_recepcion"
    )

    recientes = Encomienda.objects.filter(
        estado="entregada"
    ).select_related(
        "lote"
    ).order_by(
        "-fecha_entrega"
    )

    if busqueda:
        filtro = (
            Q(lote__numero__icontains=busqueda)
            | Q(lote__apellido_familia__icontains=busqueda)
            | Q(remitente__icontains=busqueda)
            | Q(descripcion__icontains=busqueda)
        )

        pendientes = pendientes.filter(filtro)
        recientes = recientes.filter(filtro)

    recientes = recientes[:50]

    return render(
        request,
        "core/seguridad_encomiendas.html",
        {
            "pendientes": pendientes,
            "recientes": recientes,
            "busqueda": busqueda,
        }
    )


def registrar_encomienda(request):
    if not request.user.is_authenticated:
        return redirect("home")

    if not request.user.is_staff:
        return redirect("portal")

    if request.method == "POST":
        form = EncomiendaForm(request.POST)

        if form.is_valid():
            encomienda = form.save(commit=False)
            encomienda.estado = "pendiente"
            encomienda.save()

            return redirect("seguridad_encomiendas")

    else:
        form = EncomiendaForm()

    return render(
        request,
        "core/registrar_encomienda.html",
        {
            "form": form,
        }
    )


def entregar_encomienda(request, encomienda_id):
    if not request.user.is_authenticated:
        return redirect("home")

    if not request.user.is_staff:
        return redirect("portal")

    encomienda = get_object_or_404(
        Encomienda,
        id=encomienda_id,
        estado="pendiente",
    )

    if request.method == "POST":
        form = EntregaEncomiendaForm(
            request.POST,
            lote=encomienda.lote,
        )

        if form.is_valid():
            tipo_retiro = form.cleaned_data["tipo_retiro"]

            if tipo_retiro == "familia":
                integrante = form.cleaned_data["integrante"]
                retirado_por = (
                    f"{integrante.apellido}, {integrante.nombre}"
                )
            else:
                retirado_por = form.cleaned_data["otro_nombre"].strip()

            encomienda.estado = "entregada"
            encomienda.fecha_entrega = timezone.now()
            encomienda.retirado_por = retirado_por
            encomienda.save(
                update_fields=[
                    "estado",
                    "fecha_entrega",
                    "retirado_por",
                ]
            )

            return redirect("seguridad_encomiendas")

    else:
        form = EntregaEncomiendaForm(
            lote=encomienda.lote,
        )

    return render(
        request,
        "core/entregar_encomienda.html",
        {
            "encomienda": encomienda,
            "form": form,
        }
    )


# -------------------------------------------------
# SEGURIDAD / PORTERÍA - INGRESO Y SALIDA
# -------------------------------------------------

def registrar_ingreso(request, visita_id):
    if not request.user.is_authenticated:
        return redirect("home")

    if not request.user.is_staff:
        return redirect("portal")

    visita = get_object_or_404(
        Visita,
        id=visita_id,
        estado="autorizada",
    )

    if (
        request.method == "POST"
        and visita.fecha_hora_ingreso is None
        and visita.fecha == timezone.localdate()
    ):
        visita.fecha_hora_ingreso = timezone.now()
        visita.save(update_fields=["fecha_hora_ingreso"])

    return redirect("seguridad_visitas")


def registrar_salida(request, visita_id):
    if not request.user.is_authenticated:
        return redirect("home")

    if not request.user.is_staff:
        return redirect("portal")

    visita = get_object_or_404(
        Visita,
        id=visita_id,
        estado="autorizada",
    )

    if (
        request.method == "POST"
        and visita.fecha_hora_ingreso is not None
        and visita.fecha_hora_salida is None
    ):
        visita.fecha_hora_salida = timezone.now()
        visita.save(update_fields=["fecha_hora_salida"])

    return redirect("seguridad_visitas")


# -------------------------------------------------
# LOGOUT
# -------------------------------------------------

def logout_view(request):
    logout(request)

    return redirect("home")